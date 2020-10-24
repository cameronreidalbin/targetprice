from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.image as img
import statistics, math, random
matplotlib.use('Agg')
 
book = openpyxl.load_workbook('lan/Digikey LAN Cleaned.xlsx')
sheet = book['download']
 
def getDistance(featureSet1, featureSet2):
    distance = 0
    for i in range(len(featureSet1)):
        distance += (featureSet1[i] - featureSet2[i])**2
    return distance

def quantityEstimate(quantity):
    return 4.2*quantity**(-.215)

def adjustPrice(pInitial,qInitial,qFinal):
    pFinal = pInitial * (quantityEstimate(qFinal)/quantityEstimate(qInitial))
    return pFinal

selectedFeatures = [0,0,0,0,0,0,0,0,0,0] 
 
def input(request):
    return render(request, 'lan/input.html')
 
def output(request):
    #user inputs
    selectedFeatures = [0,0,0,0,0,0,0,0,0,0] 
    b = str(request.GET['btLevel'])
    p = str(request.GET['poeLevel'])
    d = str(request.GET['discrete'])
    qc = int(request.GET['quantity'])
    if qc > 100000:
        qc = 100000    

    if b == '10':
        selectedFeatures[0] = 1
    if b == '100':
        selectedFeatures[1] = 1
    if b == '1G':
        selectedFeatures[2] = 1
    if b == '2.5G':
        selectedFeatures[3] = 1
    if b == '5G':
        selectedFeatures[4] = 1
    if b == '10G':
        selectedFeatures[5] = 1
    if p == 'poe':
        selectedFeatures[6] = 1
    if p == 'poe+':
        selectedFeatures[7] = 1
    if p == 'poe++':
        selectedFeatures[8] = 1
    if d == 'discrete':
        selectedFeatures[9] = 1
 

    #choose most similar points
    similarPoints = []
    distanceThreshold = 0
 
    while len(similarPoints) == 0:
        for rowNum in range(2,sheet.max_row+1):
            bt10 = sheet.cell(row=rowNum,column=4).value
            bt100 = sheet.cell(row=rowNum,column=5).value
            bt1G = sheet.cell(row=rowNum,column=6).value
            bt25G = sheet.cell(row=rowNum,column=7).value
            bt5G = sheet.cell(row=rowNum,column=8).value
            bt10G = sheet.cell(row=rowNum,column=9).value
            poe = sheet.cell(row=rowNum,column=10).value
            poep = sheet.cell(row=rowNum,column=11).value
            poepp = sheet.cell(row=rowNum,column=12).value
            discrete = sheet.cell(row=rowNum,column=16).value
            quantity = int(sheet.cell(row=rowNum,column=17).value)
            price = float(sheet.cell(row=rowNum,column=18).value)
            dpFeatures = [bt10,bt100,bt1G,bt25G,bt5G,bt10G,poe,poep,poepp,discrete]
            if getDistance(dpFeatures,selectedFeatures) == distanceThreshold:
                if qc/5 <= quantity <= qc*2:
                    similarPoints += [[quantity,price]]
        distanceThreshold += 1

    similarPoints.sort(key = lambda x: x[0]*x[1]*x[1])
    u = int(round(len(similarPoints)*.3,0))
    l = int(round(len(similarPoints)*.1,0))
    similarPoints = similarPoints[l:-u]
    similarPoints.sort(key=lambda x: x[1])
    similarPoints = similarPoints[:-l]


    #create points for graph
    quantities, prices = [], []
    for point in similarPoints:
        quantities += [point[0]]
        prices += [point[1]]
    priceMedian = statistics.median(prices)
    quantityMedian = statistics.median(quantities)

    graphQuantities = range( int(qc/10), int(qc*2), int(qc*(2-.1)/5) )
    quantities += graphQuantities
    for q in graphQuantities:
        p = adjustPrice(priceMedian,quantityMedian,q)*random.uniform(.75,1.25)
        prices += [p]


    #create price estimate and likely range
    priceEstimate = round(adjustPrice(priceMedian,quantityMedian,qc), 2)

    xs = range(int(qc/10),int(qc*2),100)
    eL,eM,eH = [],[],[]
    for x in xs:
        eL += [adjustPrice(priceMedian,quantityMedian,x)/1.5]
        eM += [adjustPrice(priceMedian,quantityMedian,x)]
        eH += [adjustPrice(priceMedian,quantityMedian,x)*1.5]


    #plot all that bullshit    
    plt.clf()
    plt.plot(quantities, prices,'bo')
    plt.plot(qc,priceEstimate,'r+',markersize=20)
    plt.plot(xs,eM,'b--')
    plt.fill_between(xs,eL,eH, facecolor = 'blue', alpha = .25)
    plt.title('Price of Similar Parts')
    plt.xlabel('Quantity')
    plt.ylabel('Price ($)') 

    plt.savefig('lan/static/lan/graph.png')
#    image = img.imread('lan/static/lan/' + size + '.png')
#    img.imsave('lan/static/lan/chosen.png',image)

    context = {'priceEstimate': priceEstimate, 'quantity': qc, 'bt': b, 'poe':p, 'discrete': d}
    return render(request, 'lan/output.html', context)