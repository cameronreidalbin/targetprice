from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.image as img
import statistics, random, math
matplotlib.use('Agg')
 
book = openpyxl.load_workbook('cmc/Digikey CMC Power Cleaned.xlsx')
sheet = book['download']
 
def getDistance(a1,a2,b1,b2):
    return ( (a1-a2)**2 + (b1-b2)**2 )**(1/2)

def quantityEstimate(quantity):
    return 2.686 * quantity**(-.145)

def adjustPrice(pInitial,qInitial,qFinal):
    pFinal = pInitial * (quantityEstimate(qFinal)/quantityEstimate(qInitial))
    return pFinal

normBook = openpyxl.load_workbook('cmc/CMB Data.xlsx')
normSheet = normBook['CMB - Normalized']
sizeDict = {'XS': 1680, 'S': 5544, 'M': 11262, 'L': 15960, 'XL': 28934, 'XXL': 59616}

for size in sizeDict:
    image = img.imread('staticfiles/cmc/'+size+'.png')
    img.imsave('cmc/static/cmc/'+size+'.png', image)
 
def input(request):
    return render(request, 'cmc/input.html')
 
def output(request):
    #user inputs
    lc = float(request.GET['inductance'])
    ic = float(request.GET['current'])
    qc = int(request.GET['quantity'])
    if qc > 100000:
        qc = 100000


    #find closest CMB package and get dimensions from package
    nI = (ic - .3)/(35 - .3)
    nL = (lc - .014)/(39 - .014)
    closest = 999999
    for rowNum in range(2,62):
        l = normSheet.cell(row=rowNum,column=1).value
        i = normSheet.cell(row=rowNum,column=2).value
        if getDistance(l,nL,i,nI) < closest:
            size = normSheet.cell(row=rowNum,column=7).value
            closest = getDistance(l,nL,i,nI) 
    sc = sizeDict[size]
 
    
    #choose most similar parts
    rowNums, similarPoints = [], []
    distanceThreshold = 0
 
    while len(rowNums) < 10:
        for rowNum in range(2,sheet.max_row+1):
            i = sheet.cell(row=rowNum,column=9).value
            s = sheet.cell(row=rowNum,column=10).value
            q = sheet.cell(row=rowNum,column=11).value
            p = sheet.cell(row=rowNum,column=12).value
            if (getDistance(i,ic,s,sc) < distanceThreshold) and (rowNum not in rowNums):
                if qc/5 <= q <= qc*2:
                    rowNums += [rowNum]
                    similarPoints += [[i,s,q,p]]
        distanceThreshold += 100

    similarPoints.sort(key=lambda x: x[3]*x[3]*x[2])
    u = int(len(similarPoints)*.3)
    l = int(len(similarPoints)*.1)
    similarPoints = similarPoints[l:-u]
    similarPoints.sort(key=lambda x: x[3])
    similarPoints = similarPoints[:-l]


    #create points for graph
    quantities,prices = [],[]
    for point in similarPoints:
        quantities += [point[2]]
        prices += [point[3]]
    priceMedian = statistics.median(prices)
    quantityMedian = statistics.median(quantities)

    graphQuantities = range( int(qc/10), int(qc*2), int(qc*(2-.1)/5) )
    quantities += graphQuantities
    for q in graphQuantities:
        p = adjustPrice(priceMedian,quantityMedian,q)*random.uniform(.75,1.25)
        prices += [p]


    #create price estimate and likely range
    priceEstimate = round(adjustPrice(priceMedian,quantityMedian,qc),2)

    xs = range(int(qc/10),int(qc*2),100)
    eL,eM,eH = [],[],[]
    for x in xs:
        eL += [adjustPrice(priceMedian,quantityMedian,x)/1.5]
        eM += [adjustPrice(priceMedian,quantityMedian,x)]
        eH += [adjustPrice(priceMedian,quantityMedian,x)*1.5]
 

    #plot all that bullshit    
    plt.clf()
    plt.plot(quantities, prices,'bo')
    plt.plot(qc,priceEstimate,'r+',markersize=20)
    plt.plot(xs,eM,'b--')
    plt.fill_between(xs,eL,eH, facecolor = 'blue', alpha = .25)
    plt.title('Price of Similar Parts')
    plt.xlabel('Quantity')
    plt.ylabel('Price ($)') 

    plt.savefig('cmc/static/cmc/graph.png')
    image = img.imread('cmc/static/cmc/' + size + '.png')
    img.imsave('cmc/static/cmc/chosen.png',image)
 
    context = {'priceEstimate': priceEstimate, 'quantity': qc}
    return render(request, 'cmc/output.html', context)