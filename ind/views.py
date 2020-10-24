from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.image as img
import statistics, math, random
matplotlib.use('Agg')
 
book = openpyxl.load_workbook('ind/Digikey Ind Cleaned.xlsx')
sheet = book['download']
 
def getDistance(featureSet1, featureSet2):
    distance = 0
    for i in range(len(featureSet1)):
        distance += (featureSet1[i] - featureSet2[i])**2
    return distance

def quantityEstimate(quantity):
    return 1.5*quantity**(-.15)

def adjustPrice(pInitial,qInitial,qFinal):
    pFinal = pInitial * (quantityEstimate(qFinal)/quantityEstimate(qInitial))
    return pFinal 
 
def input(request):
    return render(request, 'ind/input.html')
 
def output(request):
    #user inputs
    sc = int(request.GET['shield'])
    lc = float(request.GET['inductance'])
    ic = float(request.GET['current'])
    qc = int(request.GET['quantity'])
    if qc > 100000:
        qc = 100000


    #choose most similar parts
    rowNums,similarPoints = [],[]
    distanceThreshold = 1
    while len(similarPoints) < 10:
        for rowNum in range(2,sheet.max_row+1):
            s = sheet.cell(row=rowNum,column=7).value
            l = sheet.cell(row=rowNum,column=9).value
            i = sheet.cell(row=rowNum,column=10).value
            q = sheet.cell(row=rowNum,column=11).value
            p = sheet.cell(row=rowNum,column=12).value
            if (rowNum not in rowNums) and (s==sc):
                if ic/distanceThreshold <= i <= ic*distanceThreshold:
                    if lc/distanceThreshold <= l <= lc*distanceThreshold:
                        if qc/5 <= q <= qc*2:
                            rowNums += [rowNum]
                            similarPoints += [[q,p]]
        distanceThreshold += 1

    similarPoints.sort(key = lambda x: x[1]*x[1]*x[0])
    u = int(len(similarPoints)*.3)
    l = int(len(similarPoints)*.1)
    similarPoints = similarPoints[l:-u]
    similarPoints.sort(key=lambda x: x[1])
    similarPoints = similarPoints[:-l]


    #create points for graph
    quantities,prices = [],[]
    for point in similarPoints:
        quantities += [point[0]]
        prices += [point[1]]
    priceMedian = statistics.median(prices)
    quantityMedian = statistics.median(quantities)

    graphQuantities = range( int(qc/10), int(qc*2), int(qc*(2-.1)/5) )
    quantities += graphQuantities
    for q in graphQuantities:
        p = adjustPrice(priceMedian,quantityMedian,q)*random.uniform(.75,1.25)
        prices += [p]


    #create price estimate and likely range
    priceEstimate = round(adjustPrice(priceMedian,quantityMedian,qc), 2)

    xs = range(int(qc/10),int(qc*2),100)
    eL,eM,eH = [],[],[]
    for x in xs:
        eL += [adjustPrice(priceMedian,quantityMedian,x)/1.5]
        eM += [adjustPrice(priceMedian,quantityMedian,x)]
        eH += [adjustPrice(priceMedian,quantityMedian,x)*1.5]    
 
    plt.clf()
    plt.plot(quantities,prices,'bo')
    plt.plot(qc,priceEstimate,'r+',markersize=20)
    plt.xlabel('Quantity')
    plt.ylabel('Price ($)')
    plt.title('Price of Similar Parts')


    #plot all that bullshit    
    plt.clf()
    plt.plot(quantities, prices,'bo')
    plt.plot(qc,priceEstimate,'r+',markersize=20)
    plt.plot(xs,eM,'b--')
    plt.fill_between(xs,eL,eH, facecolor = 'blue', alpha = .25)
    plt.title('Price of Similar Parts')
    plt.xlabel('Quantity')
    plt.ylabel('Price ($)')

    plt.savefig('ind/static/ind/graph.png')
#    image = img.imread('ind/static/ind/' + size + '.png')
#    img.imsave('ind/static/ind/chosen.png',image)

    context = {'priceEstimate': priceEstimate, 'quantity': qc, 'inductance': lc, 'shield': sc, 'current': ic}
    return render(request, 'ind/output.html', context)