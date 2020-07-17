from django.shortcuts import render
from django.http import HttpResponse
from .models import Package
import openpyxl
import statistics, math, random
import matplotlib.pyplot as plt
import matplotlib.image as img
import matplotlib
matplotlib.use('Agg')

book = openpyxl.load_workbook('tr/Everyone Cleaned.xlsx')
sheet = book['Sheet1']
 
def logisticPriceEstimate(quantity, power):
    return 5/(1+math.exp(.00005*(quantity-5000))) + power/50

def getDistance(m1,m2,q1,q2):
    return ((m1-m2)**2 + (q1-q2)**2)**(1/2)
 
bobbinList = [['EE13/6/6 10-Terminal, THT, Vertical', '14.73 max.', '14.73 max.', '15.24 max.'],
              ['EE16/7/5 10-Terminal EXT, THT, Horizontal', '17.5 max.', '20 max.', '14 max.'],
              ['EE20/10/6 (EF20) 10-Terminal EXT, THT, Vertical', '22.7 max.', '14.6 max.', '25.3 max.'],
              ['EE25/13/7 (EF25) 14-Terminal EXT, THT, Horizontal', '27.05 max.', '32.25 max.', '22.86 max.'],
              ['EFD15 12-Terminal, SMT, Horizontal', '17.78 max.', '22.35 max.', '8.89 max.'],
              ['EFD20 12-Terminal EXT, SMT, Horizontal', '21.5 max.', '33.8 max.','13 max.'],
              ['EFD25 12-Terminal, SMT, Horizontal', '27.03 max.', '32.45 max.', '13.97 max.'],
              ['EFD30 12-Terminal, THT, Horizontal', '14.48 max.','33 max.','32 max.'],
              ['Small Toroidal 6-Terminal, SMT, Horizontal', '7 max.','9.1 max.', '7.8 max.'],
              ['EP7 8-Terminal, SMT, Horizontal', '9.78 max.', '9.5 max.', '10.54 max.'],
              ['EP10 8-Terminal, THT, Horizontal', '13.34 max.', '11.68 max.', '12.57 max.'],
              ['EP13 12-Terminal, SMT, Horizontal', '13.97 max.', '17.17 max.', '12.7 max.'],
              ['EPQ13 10-Terminal, SMT, Horizontal', '13.97 max.', '18.25 max.', '14.5 max.'],
              ['EPX7', '10.16 max.', '9.14 max.', '12.32 max.'],
              ['EPX9', '10.16 max.', '10.16 max.', '12.7 max.'],
              ['EPC13 10-Terminal, THT, Horizontal', '14.6 max.', '14.73 max.', '8.5 max.'],
              ['EPC17 10-Terminal, THT, Horizontal', '19 max.','18.5 max.', '12.5 max.'],
              ['EPC40 16-Terminal EXT, THT, Horizontal', '41.91 max.', '53.98 max.', '26.9 max.'],
              ['EPW15 9-Terminal EXT, SMT, Horizontal', '15.8 max.', '26.5 max.', '13.5 max.'],
              ['ER9.5 8-Terminal, SMT, Vertical', '10 max.','12.21 max.', '5.97 max.'],
              ['ER11.5 12-Terminal, SMT, Vertical', '12.95 max.', '12.85 max.', '6.35 max.'],
              ['ER14.5 12-Terminal, SMT, Vertical', '16 max.','16.8 max.', '7.62 max.'],
              ['ER28/14 12-Terminal, THT, Horizontal', '31 max.', '31 max.','25 max.'],
              ['ERL35 14-Terminal, THT, Horizontal', '36.5 max.','44 max.','28.5 max.'],
              ['ETD34 14-Terminal, THT, Horizontal', '39.6 max.', '43.18 max.', '30.48 max.'],
              ['ETD39 16-Terminal, THT, Horizontal', '49 max.','41.9 max.', '31.75 max.'],
              ['PQ2016 14-Terminal EXT, THT, Vertical', '23.9 max.', '27.58 max.', '25.2 max.'],
              ['PQ2620 12-Terminal EXT, THT, Vertical', '29.2 max.', '33 max.','30.2 max.'],
              ['PQ2625 12-Terminal EXT, THT, Vertical', '29.2 max.', '32.5 max.', '35.2 max.'],
              ['PQ3220 12-Terminal EXT, THT, Vertical', '35.56 max.', '40.1 max.','34 max.'],
              ['PQ3230 12-Terminal EXT, THT, Vertical', '35.56 max.', '38.1 max.','41.28 max.'],
              ['RM4 6-Terminal, THT, Vertical', '11.44 max.', '11.44 max.', '11.18 max.'],
              ['RM5 6-Terminal, THT, Vertical', '13.97 max.', '13.97 max.', '11.18 max.'],
              ['RM6 8-Terminal, SMT, Vertical', '20.07 max.', '21.84 max.', '13.46 max.'],
              ['RM8 10-Terminal EXT, THT, Vertical', '24.64 max.', '24.64 max.', '17.32 max.'],
              ['RM10 12-Terminal, THT, Vertical', '26.16 max.', '26.16 max.', '19.05 max.']]
 
powerPackageLookup = [('1', 0, 'Small Toroidal'),('3', 0, 'EP7'),
                      ('4', 0,'EPX7'),('6',0, 'EPX9'),('6', 0,'RM4'),
                      ('8',0, 'EP10'),('8', 0,'EPC13'),('10', 0,'RM5'),('14', 0,'EFD15'),
                      ('14', 0,'EP13'),(15,7,'EE13'),('17', 0,'EPQ13'),('19', 0,'RM6'),
                      (20,10,'EE16'),('22', 0,'EPC17'),(25,12, 'EPW15'),(38,26, 'EFD20'),
                      (42,21, 'PQ2016'),(45,30,'EE20'),('49', 0,'RM8'),
                      (68,52, 'EFD25'),(81,45, 'PQ2620'),
                      (91,56, 'RM10'),(100,70,'EE25'),(102,0, 'EFD30'),(113,80, 'PQ2625'),
                      ('120',0, 'PQ2625'),(151,84, 'PQ3220'),(261,164, 'ETD34'),
                      (311, 245, 'PQ3230'),(343,244, 'ERL35'),(437,304, 'ETD39')]
 
def index(request):
    return render(request, 'tr/index.html')
 
def input(request):
    return render(request, 'tr/input.html')
 
def output(request):
    topology = request.GET['topology']
    power = int(request.GET['power'])
    qc = int(request.GET['quantity'])
 
    #Find package by power
    if topology == 'AC/DC':
        powerPackageIndex = 1
    if topology == 'DC/DC':
        powerPackageIndex = 0
    chosen = False
    for each in powerPackageLookup:
        if int(each[powerPackageIndex]) >= power and chosen == False:
            package = each[2]
            chosen = True 
 
    #get dimensions from package    
    for each in bobbinList:
        if package in each[0]:
            l = float(each[1][0:-4])
            w = float(each[2][0:-4])
            h = float(each[3][0:-4])
    mc = l*w*h
 
    #make price estimate based on similar parts
    mm3, quantities, prices = [],[],[]
    distanceThreshold = 100

    while len(quantities) < 5:
        for rowNum in range(2,sheet.max_row+1):
            m = int(sheet.cell(row=rowNum,column=7).value)
            p = float(sheet.cell(row=rowNum,column=8).value)
            q = sheet.cell(row=rowNum,column=9).value
            newPoint = (m not in mm3) and (q not in quantities)
            if getDistance(mc,m,qc,q) < distanceThreshold and newPoint == True:
                mm3 += [m]
                quantities += [q]
                prices += [p]
        distanceThreshold += 100
        
   
    graphMin, graphMax = min(quantities), qc + max(quantities)
    graphQuantities = range(graphMin,graphMax,int(round((graphMax-graphMin)/6,0)))
    quantities += graphQuantities
    for q in graphQuantities:
        p = logisticPriceEstimate(q,power)*random.uniform(.75,1.25)
        if p <= .35:
            p = .35
        prices += [p]
 
    priceEstimate = statistics.median(prices)
    v = statistics.median(quantities)
    quantityDiscount = logisticPriceEstimate(qc,power) - logisticPriceEstimate(v,power)
    priceEstimate = round(priceEstimate+quantityDiscount,2)
 
    plt.clf()
    plt.plot(quantities, prices,'bo')
    plt.plot(qc,priceEstimate,'r+',markersize=20)
    plt.title('Distributor Pricing for Similar Parts')
    plt.xlabel('Quantity')
    plt.ylabel('Price ($)')
    plt.savefig('tr/static/tr/graph.png')
    image = img.imread(package + '.png')
    img.imsave('tr/chosen.png',image)
 
    context = {'priceEstimate': priceEstimate, 'package': package, 'quantity': qc}
    return render(request, 'tr/output.html', context)