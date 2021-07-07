from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
import requests,bs4,lxml
import os
import re
import numpy as np
import random
import matplotlib.pyplot as plt
import matplotlib.image as img
import matplotlib
matplotlib.use('Agg')



def grabPrices(pn):
    part = str(pn).upper().replace('-','')
    res = requests.get('http://www.oemstrade.com/search/' + str(part))
    soup = bs4.BeautifulSoup(res.text, 'lxml')
    points = []
   
    distributorResults = soup.select('.distributor-results')      
    for dr in distributorResults:

        volumeDistributors = ['Arrow Electronics','Avnet Americas','Avnet Europe','Avnet Asia','Future Electronics','C1C Co., Ltd.']
        onlineDistributors = ['Mouser Electronics','Digi-Key','Newark','Farnell']
        brokers = ['C1C Co., Ltd.','LCSC']
        distributorName = dr.get('data-distributorname')
        margin = 0
        if distributorName in volumeDistributors:
            margin = 1
        if distributorName in onlineDistributors:
            margin = 1.25

        if distributorName not in brokers:
            WurthPNs = ['744','742','750','760','768','691']
            if pn[0:3] in WurthPNs:
                margin += .75
       
            partNumber = dr.select('.td-part-number > a')
            if part in partNumber[0].getText().upper().replace('-',''):
                quotes = dr.select('.table-list > .multi-price')
           
                for q in quotes:
                    if q.select('.list-right')[0].get('data-basecurrency') == 'USD':
                        quantity = int(q.select('.list-left')[0].getText())
                        price = float(q.select('.list-right')[0].get('data-baseprice'))
                        adjustedPrice = price / (1+margin)
                        points += [[quantity,adjustedPrice,distributorName,price]]
   
    return points



def findBestPrices(points,chosenQuantity):
    points.sort(key = lambda x: x[0])
    currentLowest = 999
    quantities,prices,keyPrices = [],[],[]

    for p in points:
        if p[0] <= chosenQuantity*1.5:
            if p[1] <= currentLowest:
                quantities += [p[0]]
                prices += [p[1]]
                currentLowest = p[1]
                keyPrices += [ str(p[0]) + ' - ' + p[2] + ' - Distribution Price = $' + str(round(p[3],2)) + ' - Direct Price = $' + str(round(p[1],2)) ]
            else:
                quantities += [p[0]]
                prices += [currentLowest]

    return quantities,prices,keyPrices



def createPowerTrendLine(quantities,prices,chosenQuantity):    
    #y = ax^b
    #ln(y) = b*(ln(x)) + ln(a)
    logQuantities,logPrices = [],[]
    for q in quantities:
        logQuantities += [np.log(q)]  
    for p in prices:
        logPrices += [np.log(p)]
    x,y = [],[]
   
    b,lna = np.polyfit(logQuantities,logPrices,1)
    a = np.exp(lna)
    x,y = [],[]

    for i in range(1,int(chosenQuantity*1.5),int(chosenQuantity/100)):
        x += [i]
        y += [a*(i**b)]
   
    priceEstimate = round(a*(chosenQuantity**b),3)
   
    return x,y,priceEstimate



def getDatabaseCrosses(pn):
    wb = openpyxl.load_workbook('pnsearch/Inductor Cross List.xlsx')
    sheet = wb['inductors']
    pn = str(pn).upper().replace('-','')
    potentialMatches = []
    for colNum in range(1,6):
        for rowNum in range(2,sheet.max_row+1):
            cellPN = str(sheet.cell(row=rowNum,column=colNum).value).upper().replace('-','')
            if str(cellPN)[0:3] == str(pn)[0:3]:
                potentialMatches += [[cellPN,rowNum,colNum]]

    matchRequirement = 3
    newMatchList = ['']
    while len(newMatchList) > 0 and matchRequirement < len(pn):
        matchRequirement += 1
        newMatchList = []
        for match in potentialMatches:
            if match[0][:matchRequirement] == pn[:matchRequirement]:
                newMatchList += [match]
    
    if len(newMatchList) > 0:
        finalMatch = newMatchList[0]
        crosses = [pn]
        for x in range(1,5):
            cross = str(sheet.cell(row=finalMatch[1],column=x).value).upper().replace('-','')
            if cross != '-' and cross != None and cross not in crosses:
                crosses += [cross]
        return crosses
    else:
        return []




def plotStuff(pn,chosenQuantity,color,estimateLabel=False):
    pn = str(pn).upper().replace('-','')
    if pn[-7:] == ' FAMILY':
        points = grabPrices(pn[:-7])
    else:
        points = grabPrices(pn)
    quantities,prices,keyPrices = findBestPrices(points,chosenQuantity)

    if len(quantities) <= 1:
        xTrend,yTrend,priceEstimate = [],[],0
        extraX,extraY = [],[]
        page = 'pnsearch/output - error no data.html'

    if len(quantities) > 1:
        xTrend,yTrend,priceEstimate = createPowerTrendLine(quantities,prices,chosenQuantity)
        extraX, extraY = [],[]
        for i in range(len(xTrend)):
            if i%int(len(xTrend)/5) == 0 and xTrend[i]<25000:
                extraX += [xTrend[i]]
                extraY += [yTrend[i]*random.uniform(.8,1.2)]
        page = 'pnsearch/output.html'

    plt.plot(quantities,prices,c=color,ls='None',marker='o')
    plt.plot(xTrend,yTrend,c=color,ls='--',marker='None',label=pn)
    plt.plot(chosenQuantity,priceEstimate,c=color,marker='+',markerSize=20)
    if estimateLabel == True:
        plt.annotate((chosenQuantity,'$'+str(priceEstimate)),(chosenQuantity,priceEstimate),textcoords="offset points",xytext=(0,10),color=color)

    return page,priceEstimate,keyPrices

 
def input(request):
    return render(request, 'pnsearch/input.html')
 


def output(request):
    #user inputs
    pn = str(request.GET['pn']).upper().replace('-','')
    chosenQuantity = int(request.GET['quantity'])
    magneticsCrosses = int(request.GET['magneticsCrosses'])
    if chosenQuantity < 100:
        graphQuantity = 100
    else:
        graphQuantity = chosenQuantity

    plt.clf()
    plt.title('Direct Price Adjusted Quotes')
    plt.xlabel('Quantity')
    plt.ylabel('Price ($)') 
    page,pnPrice,keyPrices = plotStuff(pn,graphQuantity,'red',True)

    if magneticsCrosses == 1:
        crosses = getDatabaseCrosses(pn)
        colors = ['red','blue','green','orange','purple']
        if len(crosses) > 0:
            numCrosses = min(4,len(crosses))
            for x in range(1,numCrosses):
                plotStuff(crosses[x],chosenQuantity,colors[x])

    plt.legend(loc='best')
    plt.ylim(0)

    if len(keyPrices) > 2:
        keyPrices = keyPrices[-3:]
    else:
        keyPrices = None        

    plt.savefig('pnsearch/static/pnsearch/graph.png')
    context = {'quantity': graphQuantity, 'pn': pn, 'pnPrice': pnPrice, 'keyPrices': keyPrices}
    return render(request, page, context)



